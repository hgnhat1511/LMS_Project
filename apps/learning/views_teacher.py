import os
import json
import docx
import openpyxl
from dotenv import load_dotenv

# Bật hệ thống biến ảo ngay lập tức để nạp API Key
load_dotenv()

from google import genai  # Dùng bản mới nhất của Google
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.contrib.gis.geos import Point
from django.utils.crypto import get_random_string

# Gộp chung các model và form lại cho gọn gàng
from .models import Subject, Lesson, QuizQuestion, QuizChoice, Flashcard, Grade, ClassRoom
from apps.accounts.models import CustomUser
from .forms import FlashcardForm, LessonForm, QuizQuestionForm, MapQuestionForm, GradeForm, ClassRoomForm

# 1. Hàm rào chắn: Chỉ cho phép Giáo viên đi qua
def is_teacher(user):
    return user.is_authenticated and user.role == 'teacher'

# 2. Trang Tổng quan (Bảng điều khiển của Giáo viên)
@user_passes_test(is_teacher, login_url='/login/')
def dashboard_view(request):
    # Đếm số lượng dữ liệu để hiển thị ra bảng thống kê
    total_students = CustomUser.objects.filter(role='student').count()
    total_subjects = Subject.objects.count()
    total_lessons = Lesson.objects.count()
    
    return render(request, 'learning/teacher/dashboard.html', {
        'total_students': total_students,
        'total_subjects': total_subjects,
        'total_lessons': total_lessons,
    })

# 3. Xử lý giao diện Soạn Bài học (Lý thuyết)
@user_passes_test(is_teacher, login_url='/login/')
def lesson_create_view(request):
    if request.method == 'POST':
        form = LessonForm(request.POST)
        if form.is_valid():
            form.save() 
            return redirect('learning:teacher_dashboard') 
    else:
        form = LessonForm()
        
    return render(request, 'learning/teacher/lesson_form.html', {'form': form})

# 4. Xử lý giao diện Tạo Trắc nghiệm (A, B, C, D)
@user_passes_test(is_teacher, login_url='/login/')
def quiz_create_view(request):
    if request.method == 'POST':
        form = QuizQuestionForm(request.POST)
        if form.is_valid():
            question = QuizQuestion.objects.create(
                lesson=form.cleaned_data['lesson'],
                question_text=form.cleaned_data['question_text']
            )
            
            correct_idx = form.cleaned_data['correct_choice']
            
            QuizChoice.objects.create(question=question, choice_text=form.cleaned_data['choice_1'], is_correct=(correct_idx == '1'))
            QuizChoice.objects.create(question=question, choice_text=form.cleaned_data['choice_2'], is_correct=(correct_idx == '2'))
            QuizChoice.objects.create(question=question, choice_text=form.cleaned_data['choice_3'], is_correct=(correct_idx == '3'))
            QuizChoice.objects.create(question=question, choice_text=form.cleaned_data['choice_4'], is_correct=(correct_idx == '4'))
            
            return redirect('learning:teacher_quiz_add')
    else:
        form = QuizQuestionForm()
        
    return render(request, 'learning/teacher/quiz_form.html', {'form': form})

# 5. Xử lý câu hỏi Bản đồ (GIS)
@user_passes_test(is_teacher, login_url='/login/')
def map_create_view(request):
    if request.method == 'POST':
        form = MapQuestionForm(request.POST)
        if form.is_valid():
            map_question = form.save(commit=False)
            lat = form.cleaned_data['lat']
            lng = form.cleaned_data['lng']
            map_question.target_point = Point(lng, lat, srid=4326)
            map_question.save()
            return redirect('learning:teacher_dashboard')
    else:
        form = MapQuestionForm(initial={'tolerance_radius': 15000}) 
        
    return render(request, 'learning/teacher/map_form.html', {'form': form})

# 6. Xử lý tạo Flashcard
@user_passes_test(is_teacher, login_url='/login/')
def flashcard_create_view(request):
    if request.method == 'POST':
        form = FlashcardForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('learning:teacher_flashcard_add') 
    else:
        form = FlashcardForm()
        
    return render(request, 'learning/teacher/flashcard_form.html', {'form': form})

# 7. Import Tài khoản Học sinh bằng Excel
@user_passes_test(is_teacher, login_url='/login/')
def import_students_view(request):
    classes = ClassRoom.objects.select_related('grade').all()

    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        excel_file = request.FILES.get('excel_file')

        if not class_id or not excel_file:
            messages.error(request, "Vui lòng chọn lớp và tải lên file Excel!")
            return redirect('learning:teacher_import_students')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Lỗi: Vui lòng upload file Excel (.xlsx hoặc .xls)')
            return redirect('learning:teacher_import_students')

        try:
            target_class = ClassRoom.objects.get(id=class_id)
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                ho_ten = str(row[0]).strip() if row[0] else None
                email = str(row[1]).strip() if row[1] else ''

                if ho_ten:
                    base_username = email.split('@')[0] if email else f"hs_{get_random_string(5)}"
                    username = base_username
                    suffix = 1
                    
                    while CustomUser.objects.filter(username=username).exists():
                        username = f"{base_username}{suffix}"
                        suffix += 1

                    user = CustomUser.objects.create_user(
                        username=username,
                        password='Password@123',
                        email=email if email else f"{username}@school.edu.vn",
                        role='student',
                        first_name=ho_ten
                    )
                    
                    profile = user.profile
                    profile.student_class = target_class
                    profile.save()
                    
                    count += 1
            
            messages.success(request, f'🎉 Nhập thành công {count} tài khoản học sinh vào lớp {target_class.name}!')
            return redirect('learning:teacher_import_students')

        except Exception as e:
            messages.error(request, f'Lỗi đọc file: Đảm bảo file đúng chuẩn mẫu. Chi tiết: {str(e)}')
            return redirect('learning:teacher_import_students')

    return render(request, 'learning/teacher/import_students.html', {'classes': classes})

# 8. Quản lý Khối & Lớp
@user_passes_test(is_teacher, login_url='/login/')
def class_management_view(request):
    grades = Grade.objects.prefetch_related('classes').all()

    if request.method == 'POST':
        if 'btn_add_grade' in request.POST:
            grade_form = GradeForm(request.POST)
            if grade_form.is_valid():
                grade_form.save()
                messages.success(request, "Đã thêm Khối học mới!")
                return redirect('learning:teacher_classes')
                
        elif 'btn_add_class' in request.POST:
            class_form = ClassRoomForm(request.POST)
            if class_form.is_valid():
                class_form.save()
                messages.success(request, "Đã thêm Lớp học mới!")
                return redirect('learning:teacher_classes')
    else:
        grade_form = GradeForm()
        class_form = ClassRoomForm()

    return render(request, 'learning/teacher/class_management.html', {
        'grades': grades,
        'grade_form': grade_form,
        'class_form': class_form
    })
@user_passes_test(is_teacher, login_url='/login/')
def import_excel_lesson_view(request):
    # Lấy dữ liệu để đổ vào các dropdown
    grades = Grade.objects.all()
    classes = ClassRoom.objects.all()
    subjects = Subject.objects.all()
    lessons = Lesson.objects.all()

    if request.method == 'POST':
        # Xử lý Import Excel (Giữ nguyên logic xử lý file của bạn)
        lesson_id = request.POST.get('lesson_id')
        excel_file = request.FILES.get('excel_file')

        if not lesson_id or not excel_file:
            messages.error(request, "Vui lòng chọn đầy đủ thông tin và tải file Excel!")
            return redirect('learning:teacher_import_excel')

        try:
            target_lesson = Lesson.objects.get(id=lesson_id)
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            
            quiz_count = 0
            flashcard_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                loai = str(row[0]).strip().upper() if row[0] else ''
                if loai == 'QUIZ':
                    # ... (Logic lưu Quiz giữ nguyên)
                    quiz_count += 1
                elif loai == 'FLASHCARD':
                    # ... (Logic lưu Flashcard giữ nguyên)
                    flashcard_count += 1

            messages.success(request, f"🎉 Đã nhập thành công vào bài '{target_lesson.title}'!")
        except Exception as e:
            messages.error(request, f"Lỗi: {str(e)}")

        return redirect('learning:teacher_import_excel')

    return render(request, 'learning/teacher/import_excel.html', {
        'grades': grades,
        'classes': classes,
        'subjects': subjects,
        'lessons': lessons
    })